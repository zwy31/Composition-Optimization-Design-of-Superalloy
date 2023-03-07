import os

import numpy as np
import matplotlib.pyplot as plt
import math
import xlrd
import xlsxwriter
import random
from sklearn.model_selection import GridSearchCV, cross_val_score
from openpyxl import Workbook
from openpyxl import load_workbook
from bayes_opt import BayesianOptimization



#Disrupt data#
#wb = load_workbook('data01.xlsx')
def daluan():
    def random_data(ws):
        numbers = list(range(1, ws.max_row+1))
        random.shuffle(numbers)
        for i in numbers:
            row = ws[i]
            r = []
            for cell in row:
                r.append(cell.value)
            ws.append(r)

        ws.delete_rows(1, ws.max_row // 2)
    for ws in wb:

        random_data(ws)

   # wb.save('data01random.xlsx')
    # print("\n---save---")
    # file = "data01random.xlsx"
    # os.startfile(file)
# word = input("\n")



#Loading Data#
def load_data(TrainStartPo, TrainEndPo, TestStartPo, TestEndPo, PredStartPo, PredEndPo, FeatureNum, FilePath):

    workbook = xlrd.open_workbook(str(FilePath))
    sheet = workbook.sheet_by_name('Sheet1')
    train = []
    test = []
    pred = []

    for load_train in range(TrainStartPo-1, TrainEndPo):
        train.append(sheet.row_values(load_train))

    for load_test in range(TestStartPo-1, TestEndPo):
        test.append(sheet.row_values(load_test))

    for load_pred in range(PredStartPo-1, PredEndPo):
        pred.append(sheet.row_values(load_pred))

    TrainSet = np.array(train)
    TestSet = np.array(test)
    PredSet = np.array(pred)

    x1 , y1 = TrainSet[:,:FeatureNum] , TrainSet[:,-1]
    x2 , y2 = TestSet[:,:FeatureNum] , TestSet[:,-1]
    x3 , y3 = PredSet[:,:FeatureNum] , PredSet[:,-1]
    return x1 , y1 , x2 , y2, x3, y3

#Regression and display#
def regression_method(model, x_train, y_train, x_test, y_test, x_pred):
    model.fit(x_train, y_train)
    score = model.score(x_test, y_test)
    result = model.predict(x_test)
    pred = model.predict(x_pred)

    from sklearn import metrics
    ResidualSquare = (result - y_test)**2
    MSE = np.mean(ResidualSquare)
    MAE = np.mean(np.abs(result - y_test))
    RSS = sum(ResidualSquare)

    plt.figure(1, figsize=(16,8))

    plt.subplot(121)
    xxx = [-0.5, 1.5]
    yyy = [-0.5, 1.5]
    plt.plot(xxx, yyy, c='0', linewidth=1, linestyle=':', marker='.', alpha=0.3)
    plt.plot([400, 1200], [400, 1200])
    g1=plt.scatter(result, y_test, s=20 , c='r' , edgecolors='k' , marker='o' , alpha=0.8)
    plt.legend(handles=[g1], labels=['data'])
    plt.xlim((400, 1000))
    plt.ylim((400, 1000))
    plt.subplot(122)
    plt.plot(np.arange(len(pred)), pred, 'bo-', label='predict value')
    plt.suptitle(f'R^2 = {score}\nMSE = {MSE}\nMAE = {MAE}\nRSS = {RSS}')
    plt.show()
    return pred

#Write predicted value#
def write_predict(x_pred, pred, OutPut):
    workbook = xlsxwriter.Workbook(str(OutPut))
    worksheet = workbook.add_worksheet('Sheet1')
    for i in range(len(x_pred)):
        for j in range(len(x_pred[0])):
            worksheet.write(i , j, x_pred[i][j])
        worksheet.write(i, j+1, pred[i])
    workbook.close()
    print('done')

#Preset regression method#
from sklearn import ensemble
model_RandomForestRegressor = ensemble.RandomForestRegressor(n_estimators=100)

from sklearn import ensemble
model_AdaBoostRegressor = ensemble.AdaBoostRegressor(n_estimators=100)

from sklearn import ensemble
model_GradientBoostingRegressor = ensemble.GradientBoostingRegressor(n_estimators=100)

#Setting parameters and execution#
np.random.seed(100)
method = model_GradientBoostingRegressor #Choose a regression method#
trains = [1, 100]
tests = [1, 100]
preds = [1, 477]
features = 2 #Number of features#
path = r'C:\Users\hp\PycharmProjects\pythonProject\data01random.xlsx'
output = r'C:\Users\hp\PycharmProjects\pythonProject\data01result-1.xlsx'

x_train, y_train, x_test, y_test, x_pred, y_pred = load_data(trains[0], trains[1], tests[0], tests[1], preds[0], preds[1], features, path)   
pred = regression_method(method, x_train, y_train, x_test, y_test, x_pred)
write_predict(x_pred, pred, output)
file = "data01result-1.xlsx"

